import tkinter as tk
from tkinter import filedialog, messagebox, ttk, colorchooser
import os
from openpyxl import Workbook

import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "Arial"]
plt.rcParams["axes.unicode_minus"] = False


class ZahnerPlotter:
    COLOR_PALETTE = [
        "#e74c3c", "#3498db", "#2ecc71", "#9b59b6",
        "#f39c12", "#1abc9c", "#e67e22", "#2980b9",
        "#c0392b", "#8e44ad", "#16a085", "#d35400",
    ]

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Zahner 数据绘图工具")
        self.root.geometry("1280x720")
        self.root.minsize(960, 540)

        self.her_files: list[str] = []
        self.oer_files: list[str] = []
        self.file_settings: dict[str, dict] = {}  # path -> {color, width, label, enabled}

        self._auto_style = True  # set False during plot() to suppress redundant redraws

        self.fig, self.ax = plt.subplots(figsize=(6.5, 4.5))
        self.fig.subplots_adjust(left=0.12, right=0.95, top=0.93, bottom=0.14)

        self._build_left_panel()
        self._build_right_panel()
        self._wire_style_traces()

        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    # =================================================================
    #  LEFT PANEL — matplotlib
    # =================================================================
    def _build_left_panel(self):
        left = ttk.Frame(self.root)
        left.pack(side="left", fill="both", expand=True)
        left.grid_rowconfigure(1, weight=1)
        left.grid_columnconfigure(0, weight=1)
        left.grid_propagate(False)  # prevent children from resizing this frame

        self.canvas = FigureCanvasTkAgg(self.fig, master=left)
        self.canvas.get_tk_widget().configure(width=1, height=1)  # minimal size hint
        self.canvas.draw()

        toolbar_frame = ttk.Frame(left)
        toolbar_frame.grid(row=0, column=0, sticky="ew")
        self.toolbar = NavigationToolbar2Tk(self.canvas, toolbar_frame)
        self.toolbar.update()

        self.canvas.get_tk_widget().grid(row=1, column=0, sticky="nsew")

    # =================================================================
    #  RIGHT PANEL — controls
    # =================================================================
    def _build_right_panel(self):
        right = ttk.Frame(self.root, width=380)
        right.pack(side="right", fill="y", padx=(0, 4), pady=4)
        right.pack_propagate(False)

        canvas = tk.Canvas(right, highlightthickness=0)
        scrollbar = ttk.Scrollbar(right, orient="vertical", command=canvas.yview)
        self.control_frame = ttk.Frame(canvas)

        self.control_frame.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        canvas.create_window((0, 0), window=self.control_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def _on_enter(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)

        def _on_leave(event):
            canvas.unbind_all("<MouseWheel>")

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind("<Enter>", _on_enter)
        canvas.bind("<Leave>", _on_leave)

        # ---- section A: file loading ----
        sec_a = ttk.LabelFrame(self.control_frame, text="数据文件", padding=6)
        sec_a.pack(fill="x", padx=4, pady=(4, 2))

        her_row = ttk.Frame(sec_a)
        her_row.pack(fill="x", pady=2)
        ttk.Label(her_row, text="HER:", width=5, anchor="e").pack(side="left")
        self.btn_her = tk.Button(her_row, text="选择文件", width=8,
                                 command=self._select_her, bg="#e74c3c", fg="white",
                                 font=("Microsoft YaHei", 9))
        self.btn_her.pack(side="left", padx=(4, 0))
        self.lbl_her = ttk.Label(her_row, text="未选择", foreground="gray")
        self.lbl_her.pack(side="left", padx=6)

        oer_row = ttk.Frame(sec_a)
        oer_row.pack(fill="x", pady=2)
        ttk.Label(oer_row, text="OER:", width=5, anchor="e").pack(side="left")
        self.btn_oer = tk.Button(oer_row, text="选择文件", width=8,
                                 command=self._select_oer, bg="#3498db", fg="white",
                                 font=("Microsoft YaHei", 9))
        self.btn_oer.pack(side="left", padx=(4, 0))
        self.lbl_oer = ttk.Label(oer_row, text="未选择", foreground="gray")
        self.lbl_oer.pack(side="left", padx=6)

        # ---- section B: axis presets ----
        sec_b = ttk.LabelFrame(self.control_frame, text="坐标轴", padding=6)
        sec_b.pack(fill="x", padx=4, pady=2)

        preset_row = ttk.Frame(sec_b)
        preset_row.pack(fill="x")
        ttk.Label(preset_row, text="预设:", width=5).pack(side="left")
        self.preset_var = tk.StringVar(value="manual")
        ttk.Radiobutton(preset_row, text="HER", variable=self.preset_var,
                        value="her", command=self._on_preset).pack(side="left")
        ttk.Radiobutton(preset_row, text="OER", variable=self.preset_var,
                        value="oer", command=self._on_preset).pack(side="left")
        ttk.Radiobutton(preset_row, text="手动", variable=self.preset_var,
                        value="manual").pack(side="left")

        x_row = ttk.Frame(sec_b)
        x_row.pack(fill="x", pady=2)
        ttk.Label(x_row, text="X 轴:", width=5).pack(side="left")
        self.x_min = tk.StringVar(value="-1.623")
        self.x_max = tk.StringVar(value="-1.023")
        ttk.Entry(x_row, textvariable=self.x_min, width=8).pack(side="left", padx=2)
        ttk.Label(x_row, text="~").pack(side="left")
        ttk.Entry(x_row, textvariable=self.x_max, width=8).pack(side="left", padx=2)
        ttk.Label(x_row, text="V").pack(side="left", padx=4)

        y_row = ttk.Frame(sec_b)
        y_row.pack(fill="x", pady=2)
        ttk.Label(y_row, text="Y 轴:", width=5).pack(side="left")
        self.y_min = tk.StringVar(value="0")
        self.y_max = tk.StringVar(value="0")
        self.auto_y = tk.BooleanVar(value=True)
        ttk.Entry(y_row, textvariable=self.y_min, width=8).pack(side="left", padx=2)
        ttk.Label(y_row, text="~").pack(side="left")
        ttk.Entry(y_row, textvariable=self.y_max, width=8).pack(side="left", padx=2)
        ttk.Checkbutton(y_row, text="自动", variable=self.auto_y).pack(side="left", padx=4)

        unit_row = ttk.Frame(sec_b)
        unit_row.pack(fill="x", pady=2)
        self.use_density = tk.BooleanVar(value=False)
        ttk.Checkbutton(unit_row, text="转换为电流密度 (mA/cm²)",
                        variable=self.use_density).pack(side="left")
        ttk.Label(unit_row, text="面积:").pack(side="left", padx=(6, 2))
        self.electrode_area = tk.StringVar(value="1")
        ttk.Entry(unit_row, textvariable=self.electrode_area, width=6).pack(side="left")
        ttk.Label(unit_row, text="cm²").pack(side="left")

        # ---- section C: figure style (merged: size + grid + fonts) ----
        sec_c = ttk.LabelFrame(self.control_frame, text="图形设置", padding=6)
        sec_c.pack(fill="x", padx=4, pady=2)

        # figure size
        size_row = ttk.Frame(sec_c)
        size_row.pack(fill="x")
        ttk.Label(size_row, text="尺寸:").pack(side="left")
        self.fig_w = tk.StringVar(value="6.5")
        ttk.Spinbox(size_row, textvariable=self.fig_w, from_=2, to=20,
                    increment=0.5, width=5).pack(side="left", padx=2)
        ttk.Label(size_row, text="×").pack(side="left")
        self.fig_h = tk.StringVar(value="4.5")
        ttk.Spinbox(size_row, textvariable=self.fig_h, from_=1.5, to=15,
                    increment=0.5, width=5).pack(side="left", padx=2)
        ttk.Label(size_row, text="英寸").pack(side="left", padx=(2, 6))
        ttk.Button(size_row, text="应用尺寸", command=self._apply_figsize).pack(side="left")

        # grid
        self.show_grid = tk.BooleanVar(value=True)
        ttk.Checkbutton(size_row, text="网格", variable=self.show_grid).pack(side="left", padx=10)

        # legend & tick font
        font_row = ttk.Frame(sec_c)
        font_row.pack(fill="x", pady=(4, 2))
        ttk.Label(font_row, text="图例字体:").pack(side="left")
        self.legend_font_size = tk.StringVar(value="9")
        ttk.Spinbox(font_row, textvariable=self.legend_font_size,
                    from_=5, to=20, increment=1, width=4).pack(side="left", padx=2)
        self.legend_frame_on = tk.BooleanVar(value=False)
        ttk.Checkbutton(font_row, text="图例边框",
                        variable=self.legend_frame_on).pack(side="left", padx=8)

        font_row2 = ttk.Frame(sec_c)
        font_row2.pack(fill="x", pady=(0, 2))
        ttk.Label(font_row2, text="刻度字体:").pack(side="left")
        self.tick_font_size = tk.StringVar(value="10")
        ttk.Spinbox(font_row2, textvariable=self.tick_font_size,
                    from_=6, to=20, increment=1, width=4).pack(side="left", padx=2)
        ttk.Label(font_row2, text="(坐标轴数字)", foreground="gray").pack(side="left", padx=4)

        # ---- section D: line settings (dynamic) ----
        sec_d = ttk.LabelFrame(self.control_frame, text="曲线设置", padding=6)
        sec_d.pack(fill="both", expand=True, padx=4, pady=2)
        self.line_settings_frame = ttk.Frame(sec_d)
        self.line_settings_frame.pack(fill="both", expand=True)
        ttk.Label(self.line_settings_frame, text="加载文件后此处显示曲线设置",
                  foreground="gray").pack(pady=8)

        # ---- section E: action buttons ----
        sec_e = ttk.Frame(self.control_frame)
        sec_e.pack(fill="x", padx=4, pady=(4, 8))

        self.btn_plot = tk.Button(sec_e, text="绘图", width=10,
                                  command=self.plot, bg="#27ae60", fg="white",
                                  font=("Microsoft YaHei", 11, "bold"), height=1)
        self.btn_plot.pack(side="left", padx=2)

        self.btn_save_fig = tk.Button(sec_e, text="保存图片", width=10,
                                      command=self._save_figure, bg="#2c3e50", fg="white",
                                      font=("Microsoft YaHei", 10))
        self.btn_save_fig.pack(side="left", padx=2)

        self.btn_export = tk.Button(sec_e, text="导出 Excel", width=10,
                                    command=self._export_excel, bg="#8e44ad", fg="white",
                                    font=("Microsoft YaHei", 10))
        self.btn_export.pack(side="left", padx=2)

        self.lbl_status = ttk.Label(self.control_frame, text="", foreground="gray")
        self.lbl_status.pack(pady=(0, 4))

    # =================================================================
    #  STYLE TRACES — auto-apply when user changes style controls
    # =================================================================
    def _wire_style_traces(self):
        self.show_grid.trace_add("write", lambda *a: self._apply_style())
        self.legend_font_size.trace_add("write", lambda *a: self._apply_style())
        self.legend_frame_on.trace_add("write", lambda *a: self._apply_style())
        self.tick_font_size.trace_add("write", lambda *a: self._apply_style())
        # fig_w / fig_h use explicit "应用" button to avoid layout jitter

    def _apply_style(self):
        """Apply grid, tick font, legend settings immediately (no figure resize)."""
        if not self._auto_style:
            return
        if not self.ax.lines:
            return

        if self.show_grid.get():
            self.ax.grid(True, linestyle="--", alpha=0.4)
        else:
            self.ax.grid(False)

        try:
            tick_fs = int(self.tick_font_size.get())
            self.ax.tick_params(axis="both", labelsize=tick_fs)
        except ValueError:
            pass

        legend = self.ax.get_legend()
        if legend:
            try:
                lfs = int(self.legend_font_size.get())
                for text in legend.get_texts():
                    text.set_fontsize(lfs)
            except ValueError:
                pass
            legend.set_frame_on(self.legend_frame_on.get())

        self.canvas.draw()

    def _apply_figsize(self):
        """User clicked 'apply size' — resize figure and redraw."""
        try:
            w = float(self.fig_w.get())
            h = float(self.fig_h.get())
            self.fig.set_size_inches(w, h)
            self.canvas.draw()
        except ValueError:
            pass

    # =================================================================
    #  FILE SELECTION
    # =================================================================
    def _select_her(self):
        paths = filedialog.askopenfilenames(
            title="选择 HER 数据文件",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
        if paths:
            self.her_files = list(paths)
            self.lbl_her.config(text=f"{len(self.her_files)} 个文件", foreground="black")
            self._ensure_settings(self.her_files)
            self._refresh_line_settings()
            self._on_preset()

    def _select_oer(self):
        paths = filedialog.askopenfilenames(
            title="选择 OER 数据文件",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
        if paths:
            self.oer_files = list(paths)
            self.lbl_oer.config(text=f"{len(self.oer_files)} 个文件", foreground="black")
            self._ensure_settings(self.oer_files)
            self._refresh_line_settings()
            self._on_preset()

    def _ensure_settings(self, files):
        for f in files:
            if f not in self.file_settings:
                color = self.COLOR_PALETTE[len(self.file_settings) % len(self.COLOR_PALETTE)]
                self.file_settings[f] = {"color": color, "width": 1.5, "enabled": True}

    @property
    def all_files(self):
        return self.her_files + self.oer_files

    @property
    def active_files(self):
        return [f for f in self.all_files
                if self.file_settings.get(f, {}).get("enabled", True)]

    # =================================================================
    #  LINE SETTINGS UI (dynamic)
    # =================================================================
    def _refresh_line_settings(self):
        self._sync_line_settings()
        for w in self.line_settings_frame.winfo_children():
            w.destroy()

        all_f = self.all_files
        if not all_f:
            ttk.Label(self.line_settings_frame, text="加载文件后此处显示曲线设置",
                      foreground="gray").pack(pady=8)
            return

        hdr = ttk.Frame(self.line_settings_frame)
        hdr.pack(fill="x", pady=(0, 2))
        ttk.Label(hdr, text="✓", width=3).pack(side="left")
        ttk.Label(hdr, text="文件", width=16, anchor="w").pack(side="left")
        ttk.Label(hdr, text="颜色", width=5).pack(side="left", padx=1)
        ttk.Label(hdr, text="线宽", width=5).pack(side="left", padx=1)
        ttk.Label(hdr, text="标签").pack(side="left", padx=2)

        for i, f in enumerate(all_f):
            s = self.file_settings[f]
            row = ttk.Frame(self.line_settings_frame)
            row.pack(fill="x", pady=1)

            cb_var = tk.BooleanVar(value=s.get("enabled", True))
            cb = ttk.Checkbutton(row, variable=cb_var, width=2)
            cb.pack(side="left")
            cb._file_path = f
            cb._var = cb_var
            cb._field = "enabled"

            fname = os.path.basename(f)
            display = fname if len(fname) <= 18 else fname[:16] + "…"
            ttk.Label(row, text=display, width=18, anchor="w",
                      font=("Consolas", 9)).pack(side="left")

            color_btn = tk.Button(row, text="  ", bg=s["color"],
                                  activebackground=s["color"],
                                  width=3, relief="ridge",
                                  command=lambda p=f: self._pick_color(p))
            color_btn.pack(side="left", padx=3)

            w_var = tk.StringVar(value=str(s["width"]))
            w_spin = ttk.Spinbox(row, textvariable=w_var, from_=0.5, to=8.0,
                                 increment=0.5, width=4)
            w_spin.pack(side="left", padx=3)
            w_spin._file_path = f
            w_spin._var = w_var
            w_spin._field = "width"

            label_var = tk.StringVar(value=s.get("label", fname.rsplit(".", 1)[0]))
            label_entry = ttk.Entry(row, textvariable=label_var, width=12,
                                    font=("Microsoft YaHei", 9))
            label_entry.pack(side="left", padx=3)
            label_entry._file_path = f
            label_entry._var = label_var
            label_entry._field = "label"

    def _pick_color(self, path):
        current = self.file_settings[path]["color"]
        new_color = colorchooser.askcolor(color=current, title="选择线条颜色")
        if new_color and new_color[1]:
            self.file_settings[path]["color"] = new_color[1]
            self._refresh_line_settings()

    def _sync_line_settings(self):
        for child in self.line_settings_frame.winfo_children():
            for gc in child.winfo_children():
                if hasattr(gc, "_file_path") and hasattr(gc, "_field"):
                    path = gc._file_path
                    val = gc._var.get()
                    field = gc._field
                    if field == "width":
                        try:
                            self.file_settings[path]["width"] = float(val)
                        except ValueError:
                            pass
                    elif field == "label":
                        self.file_settings[path]["label"] = str(val)
                    elif field == "enabled":
                        self.file_settings[path]["enabled"] = bool(val)

    # =================================================================
    #  PRESETS
    # =================================================================
    def _on_preset(self):
        preset = self.preset_var.get()
        if preset == "her":
            self.x_min.set("-1.623")
            self.x_max.set("-1.023")
        elif preset == "oer":
            self.x_min.set("0.0")
            self.x_max.set("0.7")

    # =================================================================
    #  PARSE
    # =================================================================
    def parse_file(self, path):
        voltages, currents = [], []
        with open(path, "r", encoding="utf-8") as fh:
            fh.readline()
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                parts = line.split("\t")
                if len(parts) < 3:
                    continue
                voltages.append(float(parts[1]))
                currents.append(float(parts[2]))
        return voltages, currents

    # =================================================================
    #  PLOT
    # =================================================================
    def plot(self):
        all_f = self.active_files
        if not all_f:
            messagebox.showwarning("提示", "没有勾选任何曲线，请在曲线设置中勾选要绘制的文件。")
            return

        self._auto_style = False
        self._sync_line_settings()
        self.ax.clear()

        # apply figure size
        try:
            w = float(self.fig_w.get())
            h = float(self.fig_h.get())
            self.fig.set_size_inches(w, h)
        except ValueError:
            pass

        all_v = []
        all_c = []

        for f in all_f:
            try:
                v, c = self.parse_file(f)
            except Exception as e:
                messagebox.showerror("解析错误", f"无法读取文件:\n{f}\n\n{e}")
                self._auto_style = True
                return

            s = self.file_settings[f]
            label = s.get("label", os.path.basename(f).rsplit(".", 1)[0])
            area = 1.0
            if self.use_density.get():
                try:
                    area = float(self.electrode_area.get())
                except ValueError:
                    area = 1.0

            y_vals = [cur * 1000 / area for cur in c] if self.use_density.get() else c
            self.ax.plot(v, y_vals, color=s["color"], linewidth=s["width"],
                         label=label)
            all_v.extend(v)
            all_c.extend(y_vals)

        # axis limits
        try:
            x0 = float(self.x_min.get())
        except ValueError:
            x0 = min(all_v)
        try:
            x1 = float(self.x_max.get())
        except ValueError:
            x1 = max(all_v)

        if self.auto_y.get():
            margin = (max(all_c) - min(all_c)) * 0.05 or 0.1
            y0 = min(all_c) - margin
            y1 = max(all_c) + margin
            self.y_min.set(f"{y0:.4f}")
            self.y_max.set(f"{y1:.4f}")
        else:
            try:
                y0 = float(self.y_min.get())
            except ValueError:
                y0 = min(all_c)
            try:
                y1 = float(self.y_max.get())
            except ValueError:
                y1 = max(all_c)

        self.ax.set_xlim(x0, x1)
        self.ax.set_ylim(y0, y1)

        # labels
        try:
            tick_fs = int(self.tick_font_size.get())
        except ValueError:
            tick_fs = 10
        self.ax.set_xlabel("Potential (V vs. RHE)", fontsize=11)
        y_unit = "Current density (mA/cm²)" if self.use_density.get() else "Current (A)"
        self.ax.set_ylabel(y_unit, fontsize=11)
        self.ax.tick_params(axis="both", labelsize=tick_fs)

        # grid
        if self.show_grid.get():
            self.ax.grid(True, linestyle="--", alpha=0.4)
        else:
            self.ax.grid(False)

        # legend
        if len(all_f) > 1:
            try:
                lfs = int(self.legend_font_size.get())
            except ValueError:
                lfs = 9
            legend = self.ax.legend(
                fontsize=lfs,
                frameon=self.legend_frame_on.get(),
                loc="best",
            )
            if legend:
                legend.set_draggable(True)

        # title
        preset = self.preset_var.get()
        title_map = {"her": "HER LSV", "oer": "OER LSV", "manual": "LSV"}
        self.ax.set_title(title_map.get(preset, "LSV"), fontsize=13, fontweight="bold")

        self.canvas.draw()
        self._auto_style = True

        preset_label = {"her": "HER", "oer": "OER", "manual": "全部"}
        self.lbl_status.config(
            text=f"绘图完成 — {preset_label.get(preset, '全部')}，{len(all_f)} 条曲线",
            foreground="green")

    # =================================================================
    #  SAVE FIGURE
    # =================================================================
    def _save_figure(self):
        if not self.ax.lines:
            messagebox.showwarning("提示", "请先绘图再保存。")
            return
        path = filedialog.asksaveasfilename(
            title="保存图片",
            defaultextension=".png",
            filetypes=[
                ("PNG", "*.png"),
                ("SVG", "*.svg"),
                ("PDF", "*.pdf"),
                ("JPEG", "*.jpg"),
            ])
        if path:
            try:
                dpi = 300 if path.lower().endswith(".png") else 150
                self.fig.savefig(path, dpi=dpi, bbox_inches="tight")
                self.lbl_status.config(
                    text=f"已保存：{os.path.basename(path)}", foreground="green")
            except Exception as e:
                messagebox.showerror("保存失败", str(e))

    # =================================================================
    #  EXPORT EXCEL
    # =================================================================
    def _export_excel(self):
        if not self.her_files and not self.oer_files:
            messagebox.showwarning("提示", "请先选择 HER 或 OER 文件。")
            return

        out_path = filedialog.asksaveasfilename(
            title="保存 Excel 文件",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if not out_path:
            self.lbl_status.config(text="已取消", foreground="gray")
            return

        wb = Workbook()
        wb.remove(wb.active)

        if self.her_files:
            self._write_sheet(wb, "HER", self.her_files)
        if self.oer_files:
            self._write_sheet(wb, "OER", self.oer_files)

        wb.save(out_path)
        self.lbl_status.config(
            text=f"已生成：{os.path.basename(out_path)}", foreground="green")

    def _write_sheet(self, wb, sheet_name, files):
        ws = wb.create_sheet(title=sheet_name)
        parsed = []
        max_rows = 0
        for f in files:
            v, c = self.parse_file(f)
            parsed.append((os.path.basename(f), v, c))
            if len(v) > max_rows:
                max_rows = len(v)

        for i, (fname, _, _) in enumerate(parsed):
            col = i * 2 + 1
            ws.cell(row=1, column=col, value=fname)

        for i in range(len(parsed)):
            col_v = i * 2 + 1
            col_c = i * 2 + 2
            ws.cell(row=2, column=col_v, value="Voltage (V)")
            ws.cell(row=2, column=col_c, value="Current (A)")

        for row_idx in range(max_rows):
            excel_row = row_idx + 3
            for i, (_, voltages, currents) in enumerate(parsed):
                col_v = i * 2 + 1
                col_c = i * 2 + 2
                if row_idx < len(voltages):
                    ws.cell(row=excel_row, column=col_v, value=voltages[row_idx])
                    ws.cell(row=excel_row, column=col_c, value=currents[row_idx])

    def _on_close(self):
        plt.close("all")
        self.root.destroy()

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = ZahnerPlotter()
    app.run()
