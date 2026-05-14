"""Controller — wires user events to model mutations and view updates."""

import os
import tkinter as tk
from tkinter import messagebox, filedialog

from openpyxl import Workbook

from zahner_plotter.model import Model
from zahner_plotter.view import View


class Controller:
    def __init__(self, root: tk.Tk, model: Model, view: View):
        self.model = model
        self.view = view
        self._replot_pending = False

        # ── File section ─────────────────────────────────────
        view.right.file_section.set_callback(self._on_files_added)

        # ── Axis section ─────────────────────────────────────
        view.right.axis_section.set_preset_callback(self._on_preset)
        for var in [
            view.right.axis_section.preset_var,
            view.right.axis_section.auto_y,
            view.right.axis_section.use_density,
        ]:
            var.trace_add("write", lambda *a: self._schedule_replot())

        # ── Style section ────────────────────────────────────
        view.right.style_section.set_size_callback(self._on_apply_size)
        for var in [
            view.right.style_section.show_grid,
            view.right.style_section.tick_font_size,
            view.right.style_section.legend_font_size,
            view.right.style_section.legend_frame_on,
        ]:
            var.trace_add("write", lambda *a: self._schedule_replot())

        # ── Curve settings ───────────────────────────────────
        view.right.curve_settings.set_callback(self._schedule_replot)

        # ── Action buttons ───────────────────────────────────
        act = view.right.action_section
        act.btn_plot.configure(command=self._manual_replot)
        act.btn_save.configure(command=self._save_figure)
        act.btn_export.configure(command=self._export_excel)

        # ── Window close ─────────────────────────────────────
        root.protocol("WM_DELETE_WINDOW", self._on_close)

        # Defer initial draw until window is mapped and laid out
        root.after_idle(self.view.canvas.draw)

    # ══════════════════════════════════════════════════════════════
    #  Debounce
    # ══════════════════════════════════════════════════════════════

    def _schedule_replot(self):
        if self._replot_pending:
            return
        self._replot_pending = True
        self.view.root.after_idle(self._replot_now)

    def _replot_now(self):
        self._replot_pending = False
        self._do_replot()

    def _manual_replot(self):
        """Explicit '绘图' button — apply user-configured figure size."""
        self._replot_pending = False
        self._do_replot(apply_figsize=True)

    # ══════════════════════════════════════════════════════════════
    #  Event handlers
    # ══════════════════════════════════════════════════════════════

    def _on_files_added(self, paths: list[str], category: str):
        self.model.add_files(paths, category)
        self.view.right.file_section.update_label(
            category, len(self.model.files_by_category(category)))
        self.view.update_curve_ui()
        if category == "HER":
            self.model.apply_preset("her")
            self.view.right.axis_section.sync_from_model()
        elif category == "OER":
            self.model.apply_preset("oer")
            self.view.right.axis_section.sync_from_model()
        self._schedule_replot()

    def _on_preset(self, mode: str):
        self.model.apply_preset(mode)
        self.view.right.axis_section.x_min.set(self.model.x_min)
        self.view.right.axis_section.x_max.set(self.model.x_max)
        self._schedule_replot()

    def _on_apply_size(self):
        self._replot_pending = False
        self.view.right.style_section.sync_to_model()
        self.view.right.axis_section.sync_to_model()
        self.view.right.curve_settings.sync_to_model()
        try:
            w = float(self.model.fig_width)
            h = float(self.model.fig_height)
            self.view.fig.set_size_inches(w, h)
        except ValueError:
            pass
        self._do_replot()  # size already applied above

    # ══════════════════════════════════════════════════════════════
    #  Core replot
    # ══════════════════════════════════════════════════════════════

    def _do_replot(self, apply_figsize: bool = False):
        """Full replot. Sync widget→model, clear axes, redraw everything.

        apply_figsize=False: keep current figure dimensions (for auto-updates).
        apply_figsize=True: apply user-configured inch size (for '绘图' button).
        """
        self.view.right.axis_section.sync_to_model()
        self.view.right.style_section.sync_to_model()
        self.view.right.curve_settings.sync_to_model()

        active = self.model.active_files
        ax = self.view.ax
        ax.clear()

        if not active:
            self.view.canvas.draw()
            self.view.set_status("没有勾选任何曲线", "gray")
            return

        if apply_figsize:
            try:
                w = float(self.model.fig_width)
                h = float(self.model.fig_height)
                self.view.fig.set_size_inches(w, h)
            except ValueError:
                pass

        all_v, all_c = [], []
        for path in active:
            try:
                v, c = self.model.parse(path)
            except Exception as e:
                messagebox.showerror("解析错误", f"无法读取文件:\n{path}\n\n{e}")
                return

            s = self.model.files[path]
            area = 1.0
            if self.model.use_density:
                try:
                    area = float(self.model.electrode_area)
                except ValueError:
                    area = 1.0

            y_vals = [cur * 1000 / area for cur in c] if self.model.use_density else c
            ax.plot(v, y_vals, color=s["color"], linewidth=s["width"],
                    label=self.model.file_label(path), antialiased=False,
                    solid_joinstyle='miter')
            all_v.extend(v)
            all_c.extend(y_vals)

        try:
            x0 = float(self.model.x_min)
        except ValueError:
            x0 = min(all_v)
        try:
            x1 = float(self.model.x_max)
        except ValueError:
            x1 = max(all_v)

        if self.model.auto_y:
            margin = (max(all_c) - min(all_c)) * 0.05 or 0.1
            y0 = min(all_c) - margin
            y1 = max(all_c) + margin
            self.model.y_min = f"{y0:.4f}"
            self.model.y_max = f"{y1:.4f}"
        else:
            try:
                y0 = float(self.model.y_min)
            except ValueError:
                y0 = min(all_c)
            try:
                y1 = float(self.model.y_max)
            except ValueError:
                y1 = max(all_c)

        ax.set_xlim(x0, x1)
        ax.set_ylim(y0, y1)

        try:
            tick_fs = int(self.model.tick_font_size)
        except ValueError:
            tick_fs = 10
        ax.set_xlabel("Potential (V vs. RHE)", fontsize=11)
        y_unit = "Current density (mA/cm²)" if self.model.use_density else "Current (A)"
        ax.set_ylabel(y_unit, fontsize=11)
        ax.tick_params(axis="both", labelsize=tick_fs, direction='in', width=1, which='major')
        ax.tick_params(axis="both", direction='in', width=0.5, which='minor')
        ax.minorticks_on()

        preset = self.model.preset_mode
        title_map = {"her": "HER LSV", "oer": "OER LSV", "manual": "LSV"}
        ax.set_title(title_map.get(preset, "LSV"), fontsize=13, fontweight="bold")

        if self.model.show_grid:
            ax.grid(True, linestyle="--", alpha=0.4)

        if len(active) > 1:
            try:
                lfs = int(self.model.legend_font_size)
            except ValueError:
                lfs = 9
            legend = ax.legend(fontsize=lfs, frameon=self.model.legend_frame_on, loc="best")
            if legend:
                legend.set_draggable(True)

        self.view.canvas.draw()

        preset_label = {"her": "HER", "oer": "OER", "manual": "全部"}
        self.view.set_status(
            f"绘图完成 — {preset_label.get(preset, '全部')}，{len(active)} 条曲线")

    # ══════════════════════════════════════════════════════════════
    #  Save / Export
    # ══════════════════════════════════════════════════════════════

    def _save_figure(self):
        if not self.view.ax.lines:
            messagebox.showwarning("提示", "请先绘图再保存。")
            return
        path = filedialog.asksaveasfilename(
            title="保存图片", defaultextension=".png",
            filetypes=[("PNG", "*.png"), ("SVG", "*.svg"), ("PDF", "*.pdf"), ("JPEG", "*.jpg")])
        if path:
            try:
                dpi = 400 if path.lower().endswith(".png") else 200
                self.view.fig.savefig(path, dpi=dpi, bbox_inches="tight")
                self.view.set_status(f"已保存：{os.path.basename(path)}")
            except Exception as e:
                messagebox.showerror("保存失败", str(e))

    def _export_excel(self):
        her_files = self.model.files_by_category("HER")
        oer_files = self.model.files_by_category("OER")
        if not her_files and not oer_files:
            messagebox.showwarning("提示", "请先选择文件。")
            return

        out_path = filedialog.asksaveasfilename(
            title="保存 Excel 文件", defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if not out_path:
            return

        wb = Workbook()
        wb.remove(wb.active)
        if her_files:
            self._write_sheet(wb, "HER", her_files)
        if oer_files:
            self._write_sheet(wb, "OER", oer_files)
        wb.save(out_path)
        self.view.set_status(f"已生成：{os.path.basename(out_path)}")

    def _write_sheet(self, wb, sheet_name, files):
        ws = wb.create_sheet(title=sheet_name)
        parsed = []
        max_rows = 0
        for f in files:
            v, c = self.model.parse(f)
            parsed.append((os.path.basename(f), v, c))
            if len(v) > max_rows:
                max_rows = len(v)

        for i, (fname, _, _) in enumerate(parsed):
            col = i * 2 + 1
            ws.cell(row=1, column=col, value=fname)
        for i in range(len(parsed)):
            col_v = i * 2 + 1
            col_c = i * 2 + 2
            ws.cell(row=2, column=col_v, value="Voltage (V)")
            ws.cell(row=2, column=col_c, value="Current (A)")
        for row_idx in range(max_rows):
            excel_row = row_idx + 3
            for i, (_, voltages, currents) in enumerate(parsed):
                col_v = i * 2 + 1
                col_c = i * 2 + 2
                if row_idx < len(voltages):
                    ws.cell(row=excel_row, column=col_v, value=voltages[row_idx])
                    ws.cell(row=excel_row, column=col_c, value=currents[row_idx])

    def _on_close(self):
        import matplotlib.pyplot as plt
        plt.close("all")
        self.view.root.destroy()

    def run(self):
        self.view.root.mainloop()
